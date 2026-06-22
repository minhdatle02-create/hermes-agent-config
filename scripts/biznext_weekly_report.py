import urllib.request
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

API_URL = "https://leadapi.biznext.vn/api/public/leads"
API_KEY = "ffVdNG9cc5H9LoSJvrG1BACJVlpEuTw8"
REPORT_YEAR = 2026
REPORT_MONTH = 5


def fetch_leads():
    req = urllib.request.Request(
        API_URL,
        headers={"X-API-KEY": API_KEY, "Accept": "application/json"},
        method="GET",
    )
    ctx = None
    if API_URL.startswith("https"):
        import ssl

        ctx = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
        raw = resp.read().decode("utf-8", errors="replace")
    return json.loads(raw)


def style_header(ws, row):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def parse_date(value):
    if not value:
        return None
    value = str(value).strip()
    if not value:
        return None
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def in_report_month(item):
    d = parse_date(item.get("createdAt") or item.get("contractDate"))
    if not d:
        return False
    return d.year == REPORT_YEAR and d.month == REPORT_MONTH


def is_contract(item):
    return (item.get("status") or "").strip().lower() in ("contract", "closed")


def is_following(item):
    status = (item.get("status") or "").strip().lower()
    expected = parse_date(item.get("expectedContractDate"))
    return status in ("hot", "warm") or expected is not None


def fmt_number(value):
    try:
        return float(value)
    except Exception:
        return 0.0


def build_workbook(leads):
    wb = Workbook()
    wb.remove(wb.active)

    leads_cur = [x for x in leads if in_report_month(x)]
    sales = sorted({x.get("sale") for x in leads_cur if x.get("sale")})

    # Tổng quan
    ws1 = wb.create_sheet("Tong_quan")
    ws1.append(["KPI", "Gia tri"])
    ws1.append(["Tong so lead trong thang", len(leads_cur)])
    ws1.append(
        [
            "Tong hop dong trong thang",
            sum(1 for x in leads_cur if is_contract(x)),
        ]
    )
    ws1.append(
        [
            "Tong doanh thu dich vu trong thang (VND)",
            sum(fmt_number(x.get("revenue")) for x in leads_cur),
        ]
    )
    ws1.append(
        [
            "So co hoi dang follow",
            sum(1 for x in leads_cur if is_following(x)),
        ]
    )
    ws1.append(
        [
            "Ngay bao cao",
            datetime.today().strftime("%d/%m/%Y"),
        ]
    )
    ws1.append(["Kỳ báo cáo", f"Tháng {REPORT_MONTH:02d}/{REPORT_YEAR}"])
    style_header(ws1, 1)
    add_border(ws1, 1, ws1.max_row, 1, ws1.max_column)
    ws1.column_dimensions["A"].width = 34
    ws1.column_dimensions["B"].width = 24

    # Lead theo sale
    ws2 = wb.create_sheet("Lead_theo_sale")
    ws2.append(
        [
            "Sale",
            "So lead tao trong thang",
            "Hop dong thang nay",
            "Co hoi dang follow",
            "Doanh thu tu hop dong (VND)",
        ]
    )
    for s in sales:
        month_leads = [x for x in leads_cur if x.get("sale") == s]
        contracts = [x for x in month_leads if is_contract(x)]
        follow = [x for x in month_leads if is_following(x)]
        ws2.append(
            [
                s,
                len(month_leads),
                len(contracts),
                len(follow),
                sum(fmt_number(x.get("revenue")) for x in contracts),
            ]
        )
    style_header(ws2, 1)
    add_border(ws2, 1, ws2.max_row, 1, ws2.max_column)
    for col in ("A", "B", "C", "D", "E"):
        ws2.column_dimensions[col].width = 24

    # Co hoi theo trang thai
    ws3 = wb.create_sheet("Co_hoi_trang_thai")
    ws3.append(["Trang thai", "So luong"])
    status_map = {}
    for x in leads_cur:
        k = (x.get("status") or "Unknown").strip()
        status_map[k] = status_map.get(k, 0) + 1
    for k, v in status_map.items():
        ws3.append([k, v])
    style_header(ws3, 1)
    add_border(ws3, 1, ws3.max_row, 1, ws3.max_column)
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 16

    # Doanh thu theo dich vu
    ws4 = wb.create_sheet("Doanh_thu_dich_vu")
    ws4.append(["Dich vu", "So hop dong", "Doanh thu (VND)"])
    service_map = {}
    for x in leads_cur:
        if is_contract(x):
            svc = (x.get("service") or "Khac").strip() or "Khac"
            if svc not in service_map:
                service_map[svc] = {"count": 0, "revenue": 0.0}
            service_map[svc]["count"] += 1
            service_map[svc]["revenue"] += fmt_number(x.get("revenue"))
    for svc, val in service_map.items():
        ws4.append([svc, val["count"], val["revenue"]])
    style_header(ws4, 1)
    add_border(ws4, 1, ws4.max_row, 1, ws4.max_column)
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 22

    # Follow list
    ws5 = wb.create_sheet("Follow_list")
    ws5.append(
        [
            "ID",
            "Ten",
            "Sale",
            "Trang thai",
            "Dich vu",
            "Ngay tao",
            "Ngay du kien HD",
            "Ghi chu",
        ]
    )
    follows = [x for x in leads_cur if is_following(x)]
    for x in follows:
        ws5.append(
            [
                x.get("id"),
                x.get("name"),
                x.get("sale"),
                x.get("status"),
                x.get("service"),
                x.get("createdAt"),
                x.get("expectedContractDate"),
                str(x.get("notes") or "")[:200],
            ]
        )
    style_header(ws5, 1)
    add_border(ws5, 1, ws5.max_row, 1, ws5.max_column)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws5.column_dimensions[col].width = 24

    out_dir = os.path.join(
        os.environ.get("USERPROFILE", os.path.expanduser("~")),
        "AppData",
        "Local",
        "hermes",
        "cron",
        "output",
    )
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(
        out_dir, f"BizNext_Weekly_{REPORT_YEAR}{REPORT_MONTH:02d}.xlsx"
    )
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    print("Dang lay leads tu CRM...", flush=True)
    leads = fetch_leads()
    print(f"Nhan duoc {len(leads)} leads")
    out = build_workbook(leads)
    print(out)
