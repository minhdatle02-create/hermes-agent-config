from __future__ import annotations

import json
import os
import urllib.request
from datetime import datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


API_URL = os.environ.get(
    "BIZNEXT_API_URL",
    "https://leadapi.biznext.vn/api/public/leads",
)
API_KEY = os.environ.get("BIZNEXT_API_KEY", "")
OUTPUT_DIR = os.path.join(os.environ.get("USERPROFILE", os.path.expanduser("~")), "AppData", "Local", "hermes", "cron", "output")


def fetch_leads() -> list[dict[str, Any]]:
    headers = {"Accept": "application/json"}
    if API_KEY:
        headers["X-API-KEY"] = API_KEY
    req = urllib.request.Request(API_URL, headers=headers, method="GET")
    ctx = None
    if API_URL.startswith("https"):
        import ssl
        ctx = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
        body = resp.read().decode("utf-8", errors="replace")
        print(f"Fetch ok status={resp.status} bytes={len(body)}", flush=True)
    return json.loads(body)


def style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def add_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def _date_in(date_str: str, start_date: datetime, end_date: datetime) -> bool:
    d = (date_str or "").strip()
    if not d:
        return False
    try:
        dt = datetime.strptime(d[:10], "%Y-%m-%d")
        return start_date <= dt <= end_date
    except Exception:
        return False


def build_weekly_workbook(leads: list[dict[str, Any]], start_date: datetime, end_date: datetime) -> str:
    status_ok = ("contract", "closed")

    def in_window(x: dict[str, Any]) -> bool:
        cd = x.get("contractDate") or ""
        st = str(x.get("status") or "").strip().lower()
        return _date_in(cd, start_date, end_date) and st in status_ok

    week_leads = [x for x in leads if in_window(x)]
    sales = sorted({x.get("sale") for x in week_leads if x.get("sale")})
    label = f"{start_date:%d/%m/%Y} - {end_date:%d/%m/%Y}"
    today_label = datetime.today().strftime("%d/%m/%Y")

    # Summary
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Tong_quan")
    ws.append(["KPI", "Gia tri"])
    ws.append(["Khoang thoi gian", label])
    ws.append(["So hop dong xac nhan", len(week_leads)])
    ws.append(["Doanh thu hop dong (VND)", sum(float(x.get("revenue") or 0) for x in week_leads)])
    ws.append(["So sale co hop dong", len(sales)])
    ws.append(["Ngay tao bao cao", today_label])
    style_header(ws, 1)
    add_border(ws, 1, ws.max_row, 1, ws.max_column)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26

    # Sales breakdown
    ws = wb.create_sheet("Lead_theo_sale")
    ws.append(["Sale", "Hop dong", "Doanh thu hop dong (VND)", "Doanh thu TB/hop dong (VND)"])
    for s in sales:
        items = [x for x in week_leads if x.get("sale") == s]
        ws.append([
            s,
            len(items),
            sum(float(x.get("revenue") or 0) for x in items),
            round(sum(float(x.get("revenue") or 0) for x in items) / len(items), 2) if items else 0,
        ])
    style_header(ws, 1)
    add_border(ws, 1, ws.max_row, 1, ws.max_column)
    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 26

    # Service breakdown
    ws = wb.create_sheet("Doanh_thu_dich_vu")
    ws.append(["Dich vu", "So hop dong", "Doanh thu (VND)"])
    svc_map: dict[str, dict[str, float]] = {}
    for x in week_leads:
        svc = (x.get("service") or "Khac").strip() or "Khac"
        svc_map.setdefault(svc, {"count": 0, "revenue": 0.0})
        svc_map[svc]["count"] += 1
        svc_map[svc]["revenue"] += float(x.get("revenue") or 0)
    for svc, val in svc_map.items():
        ws.append([svc, int(val["count"]), val["revenue"]])
    style_header(ws, 1)
    add_border(ws, 1, ws.max_row, 1, ws.max_column)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24

    # Contract detail
    ws = wb.create_sheet("Hop_dong")
    ws.append(["ID", "Ten", "Sale", "Trang thai", "Dich vu", "Ngay HD", "Ngay du kien HD", "Doanh thu (VND)", "Ghi chu"])
    for x in week_leads:
        ws.append([
            x.get("id"),
            x.get("name"),
            x.get("sale"),
            x.get("status"),
            x.get("service"),
            x.get("contractDate"),
            x.get("expectedContractDate"),
            float(x.get("revenue") or 0),
            str(x.get("notes") or "")[:240],
        ])
    style_header(ws, 1)
    add_border(ws, 1, ws.max_row, 1, ws.max_column)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I"):
        ws.column_dimensions[col].width = 24

    # All new leads in window (ensure-ledger = createdAt in window) as optional monitor sheet
    created_window = [x for x in leads if _date_in(x.get("createdAt") or "", start_date, end_date)]
    ws = wb.create_sheet("Lead_moi_createdAt")
    ws.append(["ID", "Ten", "Sale", "Trang thai", "Dich vu", "Ngay tao", "Ngay HD", "Doanh thu (VND)"])
    for x in created_window:
        ws.append([
            x.get("id"),
            x.get("name"),
            x.get("sale"),
            x.get("status"),
            x.get("service"),
            x.get("createdAt"),
            x.get("contractDate"),
            float(x.get("revenue") or 0),
        ])
    style_header(ws, 1)
    add_border(ws, 1, ws.max_row, 1, ws.max_column)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 24

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = os.path.join(OUTPUT_DIR, f"BizNext_weekly_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx")
    wb.save(out)
    return out


def main() -> None:
    today = datetime.today().date()
    # Start from Monday of last week through Sunday
    start_date = today - timedelta(days=today.weekday() + 7)
    end_date = start_date + timedelta(days=6)
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    print(f"Target window: {start_dt:%Y-%m-%d} -> {end_dt:%Y-%m-%d}", flush=True)
    leads = fetch_leads()
    print(f"Fetched total leads: {len(leads)}", flush=True)

    out = build_weekly_workbook(leads, start_dt, end_dt)
    print(f"Report: {out}", flush=True)


if __name__ == "__main__":
    main()
