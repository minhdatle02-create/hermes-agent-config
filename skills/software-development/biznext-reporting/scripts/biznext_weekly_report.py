from __future__ import annotations

import json
import os
import urllib.request
from datetime import datetime, timedelta
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except Exception:
    Workbook = None

API_URL = os.environ.get(
    "BIZNEXT_API_URL",
    "https://leadapi.biznext.vn/api/public/leads",
)

API_KEY = ""
_env_path = os.path.join(os.environ.get("USERPROFILE", os.path.expanduser("~")), "AppData", "Local", "hermes", ".env")
if os.path.exists(_env_path):
    with open(_env_path, "r", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            if line.startswith("BIZNEXT_API_KEY="):
                API_KEY = line.split("=", 1)[1]
                break

OUTPUT_DIR = os.path.join(os.environ.get("USERPROFILE", os.path.expanduser("~")), "AppData", "Local", "hermes", "cron", "output")


def fetch_leads() -> list[dict[str, Any]]:
    headers = {"Accept": "application/json"}
    if API_KEY:
        headers["X-API-KEY"] = API_KEY
    req = urllib.request.Request(API_URL, headers=headers, method="GET")
    ctx = None
    if API_URL.startswith("https"):
        import ssl
        ctx = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
        body = resp.read().decode("utf-8", errors="replace")
        print(f"Fetch ok status={resp.status} bytes={len(body)}", flush=True)
    return json.loads(body)


def _parse_date(date_str: str) -> datetime | None:
    d = (date_str or "").strip()
    if not d:
        return None
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d")
    except Exception:
        return None


def build_text_briefing(leads: list[dict[str, Any]], start_date: datetime, end_date: datetime) -> tuple[str, str]:
    status_ok = ("contract", "closed")
    week_contracts = [
        x for x in leads
        if (
            (dt := _parse_date(x.get("contractDate") or "")) is not None
            and start_date <= dt <= end_date
            and str(x.get("status") or "").strip().lower() in status_ok
        )
    ]
    total = len(leads)
    confirmed = len(week_contracts)
    revenue = sum(float(x.get("revenue") or 0) for x in week_contracts)

    follow = [
        x for x in leads
        if str(x.get("status") or "").strip().lower() not in status_ok
    ]
    follow_active = sorted(follow, key=lambda x: _parse_date(x.get("expectedContractDate") or "") or datetime.min, reverse=True)[:20]

    source_counts: dict[str, int] = {}
    for x in leads:
        src = (x.get("source") or "Không rõ").strip() or "Không rõ"
        source_counts[src] = source_counts.get(src, 0) + 1

    lines = []
    lines.append(f"Báo cáo nhanh tuần trước ({start_date:%d/%m} – {end_date:%d/%m}):")
    lines.append("")
    lines.append("Tổng quan")
    lines.append(f"- Dataset: {total} leads")
    lines.append(f"- Confirmed contracts trong tuần: {confirmed}")
    lines.append(f"- Doanh thu hợp đồng: {int(revenue):,}đ")
    lines.append(f"- Follow đang active: {len(follow)}")
    lines.append("")
    lines.append("Cơ hội mới")
    if week_contracts:
        for i, x in enumerate(week_contracts, 1):
            lines.append(f"{i}. {x.get('id')} — {x.get('name')} — {x.get('sale')}")
            lines.append(f"   - Trạng thái: {x.get('status')}")
            lines.append(f"   - Dịch vụ: {x.get('service')}")
            lines.append(f"   - Nguồn: {x.get('source')}")
            lines.append(f"   - Ngày HD: {x.get('contractDate')}")
            lines.append(f"   - Doanh thu: {int(float(x.get('revenue') or 0)):,}đ")
            lines.append(f"   - Ghi chú: {x.get('notes')}")
    else:
        lines.append("- Không có cơ hội mới đã chốt tuần trước.")
    lines.append("")
    lines.append("Nguồn cơ hội (dataset toàn bộ)")
    for src, count in sorted(source_counts.items(), key=lambda kv: kv[1], reverse=True)[:10]:
        pct = round(count / total * 100, 1) if total else 0.0
        lines.append(f"- {src}: {count} ({pct}%)")
    lines.append("")
    lines.append("Follow list cần đẩy")
    if follow_active:
        for x in follow_active[:10]:
            lines.append(
                f"- {x.get('id')}: {x.get('name')} — {x.get('sale')} — {x.get('status')} — {x.get('service')} — {x.get('source')} — HĐ dự kiến: {x.get('expectedContractDate')} — {x.get('notes')}"
            )
    else:
        lines.append("- Không có follow.")
    return "\n".join(lines), ""


def build_weekly_workbook(leads: list[dict[str, Any]], start_date: datetime, end_date: datetime) -> str:
    if Workbook is None:
        raise RuntimeError("openpyxl không khả dụng")
    status_ok = ("contract", "closed")
    week_contracts = [
        x for x in leads
        if (
            (dt := _parse_date(x.get("contractDate") or "")) is not None
            and start_date <= dt <= end_date
            and str(x.get("status") or "").strip().lower() in status_ok
        )
    ]
    sales = sorted({x.get("sale") for x in week_contracts if x.get("sale")})
    label = f"{start_date:%d/%m/%Y} - {end_date:%d/%m/%Y}"
    today_label = datetime.today().strftime("%d/%m/%Y")

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Tong_quat")
    ws.append(["KPI", "Gia tri"])
    ws.append(["Khoang thoi gian", label])
    ws.append(["So hop dong", len(week_contracts)])
    ws.append(["Doanh thu hop dong (VND)", sum(float(x.get("revenue") or 0) for x in week_contracts)])
    ws.append(["So sale", len(sales)])
    ws.append(["Ngay tao bao cao", today_label])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26

    ws = wb.create_sheet("Hop_dong_chot")
    ws.append(["ID", "Ten", "Sale", "Trang thai", "Dich vu", "Nguon", "Ngay HD", "Ngay du kien HD", "Doanh thu (VND)", "Ghi chu"])
    for x in week_contracts:
        ws.append([
            x.get("id"), x.get("name"), x.get("sale"), x.get("status"), x.get("service"),
            x.get("source"), x.get("contractDate"), x.get("expectedContractDate"),
            float(x.get("revenue") or 0), str(x.get("notes") or "")[:240],
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[col].width = 24

    ws = wb.create_sheet("Doanh_thu_dich_vu")
    ws.append(["Dich vu", "So hop dong", "Doanh thu (VND)"])
    svc_map: dict[str, dict[str, float]] = {}
    for x in week_contracts:
        svc = (x.get("service") or "Khac").strip() or "Khac"
        svc_map.setdefault(svc, {"count": 0, "revenue": 0.0})
        svc_map[svc]["count"] += 1
        svc_map[svc]["revenue"] += float(x.get("revenue") or 0)
    for svc, val in svc_map.items():
        ws.append([svc, int(val["count"]), val["revenue"]])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24

    ws = wb.create_sheet("Lead_theo_sale")
    ws.append(["Sale", "Hop dong", "Doanh thu hop dong (VND)", "Doanh thu TB/hop dong (VND)"])
    for s in sales:
        items = [x for x in week_contracts if x.get("sale") == s]
        revenue = sum(float(x.get("revenue") or 0) for x in items)
        ws.append([s, len(items), revenue, round(revenue / len(items), 2) if items else 0])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 26

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = os.path.join(OUTPUT_DIR, f"BizNext_weekly_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx")
    wb.save(out)
    return out


def main() -> None:
    today = datetime.today().date()
    start_date = today - timedelta(days=today.weekday() + 7)
    end_date = start_date + timedelta(days=6)
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    print(f"Target window: {start_dt:%Y-%m-%d} -> {end_dt:%Y-%m-%d}", flush=True)
    leads = fetch_leads()
    print(f"Fetched total leads: {len(leads)}", flush=True)

    try:
        out = build_weekly_workbook(leads, start_dt, end_dt)
        print(f"Report: {out}", flush=True)
    except Exception as e:
        print(f"Workbook generation failed: {e}", flush=True)

    briefing, _ = build_text_briefing(leads, start_dt, end_dt)
    print("\n---BRIEFING---")
    print(briefing)


if __name__ == "__main__":
    main()
