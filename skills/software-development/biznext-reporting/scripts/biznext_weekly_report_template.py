from __future__ import annotations

import json
import os
import urllib.request
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


API_URL = os.environ.get(
    "BIZNEXT_API_URL",
    "https://leadapi.biznext.vn/api/public/leads",
)
API_KEY = os.environ.get("BIZNEXT_API_KEY", "")  # configured via hermes config
OUTPUT_DIR = os.path.join(os.environ.get("USERPROFILE", os.path.expanduser("~")), "AppData", "Local", "hermes", "cron", "output")


def fetch_leads() -> list[dict[str, Any]]:
    headers = {"Accept": "application/json"}
    if API_KEY:
        headers["X-API-KEY"] = API_KEY
    req = urllib.request.Request(API_URL, headers=headers, method="GET")
    ctx = None
    if API_URL.startswith("https"):
        import ssl
        ctx = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
        body = resp.read().decode("utf-8", errors="replace")
    return json.loads(body)


def style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def add_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def month_test(item: dict[str, Any], year: int, month: int) -> bool:
    d = (item.get("createdAt") or item.get("contractDate") or "").strip()
    if not d:
        return False
    try:
        dt = datetime.strptime(d[:10], "%Y-%m-%d")
        return dt.year == year and dt.month == month
    except Exception:
        return False


def build_workbook(leads: list[dict[str, Any]], year: int, month: int) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    leads_cur = [x for x in leads if month_test(x, year, month)]
    sales = sorted({x.get("sale") for x in leads_cur if x.get("sale")})
    label = f"{month:02d}/{year}"
    today_label = datetime.today().strftime("%d/%m/%Y")

    # --- Summary ---
    ws = wb.create_sheet("Tong_quan")
    ws.append(["KPI", "Gia tri"])
    status_ok = ("contract", "closed")
    ws.append(["Thang bao cao", label])
    ws.append(["So luong lead trong thang", len(leads_cur)])
    ws.append(["So hop dong trong thang", sum(1 for x in leads_cur if str(x.get("contractDate") or "").startswith(f"{year}-{month:02d}") and str(x.get("status") or "").lower() in status_ok)])
    ws.append(["Doanh thu dich vu trong thang (VND)", sum(float(x.get("revenue") or 0) for x in leads_cur)])
    ws.append(["So lead dang follow", sum(1 for x in leads_cur if str(x.get("status") or "").lower() in ("hot", "warm") or str(x.get("expectedContractDate") or "").strip() not in ("", ""))])
    ws.append(["Ngay tao bao cao", today_label])
    style_header(ws, 1); add_border(ws, 1, ws.max_row, 1, ws.max_column)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26

    # --- Sales breakdown ---
    ws = wb.create_sheet("Lead_theo_sale")
    ws.append(["Sale", "So lead trong thang", "Hop dong", "Dang follow", "Doanh thu hop dong (VND)"])
    status_ok = ("contract", "closed")
    for s in sales:
        items = [x for x in leads_cur if x.get("sale") == s]
        contracts = [x for x in items if str(x.get("status") or "").lower() in status_ok]
        follow = [x for x in items if str(x.get("status") or "").lower() in ("hot", "warm") or str(x.get("expectedContractDate") or "").strip() not in ("")]
        ws.append([s, len(items), len(contracts), len(follow), sum(float(x.get("revenue") or 0) for x in contracts)])
    style_header(ws, 1); add_border(ws, 1, ws.max_row, 1, ws.max_column)
    for col in ("A", "B", "C", "D", "E"):
        ws.column_dimensions[col].width = 24

    # --- Source breakdown ---
    ws = wb.create_sheet("Nguon_cau_hoi")
    ws.append(["Nguon", "So luong"])
    source_map: dict[str, int] = {}
    for x in leads_cur:
        src = (x.get("source") or "Khong ro").strip() or "Khong ro"
        source_map[src] = source_map.get(src, 0) + 1
    for k, v in source_map.items():
        ws.append([k, v])
    style_header(ws, 1); add_border(ws, 1, ws.max_row, 1, ws.max_column)
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 16

    # --- Status ---
    ws = wb.create_sheet("Co_hoi_trang_thai")
    ws.append(["Trang thai", "So luong"])
    counts: dict[str, int] = {}
    for x in leads_cur:
        k = (x.get("status") or "Unknown").strip() or "Unknown"
        counts[k] = counts.get(k, 0) + 1
    for k, v in counts.items():
        ws.append([k, v])
    style_header(ws, 1); add_border(ws, 1, ws.max_row, 1, ws.max_column)
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 16

    # --- Revenue by service ---
    ws = wb.create_sheet("Doanh_thu_dich_vu")
    ws.append(["Dich vu", "So hop dong", "Doanh thu (VND)"])
    service_map: dict[str, dict[str, float]] = {}
    for x in leads_cur:
        if str(x.get("status") or "").lower() in status_ok:
            svc = (x.get("service") or "Khac").strip() or "Khac"
            service_map.setdefault(svc, {"count": 0, "revenue": 0.0})
            service_map[svc]["count"] += 1
            service_map[svc]["revenue"] += float(x.get("revenue") or 0)
    for svc, val in service_map.items():
        ws.append([svc, val["count"], val["revenue"]])
    style_header(ws, 1); add_border(ws, 1, ws.max_row, 1, ws.max_column)
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 16; ws.column_dimensions["C"].width = 22

    # --- Follow list ---
    ws = wb.create_sheet("Follow_list")
    ws.append(["ID", "Ten", "Sale", "Trang thai", "Dich vu", "Ngay tao", "Ngay du kien HD", "Ghi chu"])
    follows = [x for x in leads_cur if (
        str(x.get("status") or "").lower() in status_ok
        or str(x.get("status") or "").lower() in ("hot", "warm")
        or str(x.get("expectedContractDate") or "").strip() not in ("")
    )]
    for x in follows:
        ws.append([
            x.get("id"), x.get("name"), x.get("sale"), x.get("status"), x.get("service"),
            x.get("createdAt"), x.get("expectedContractDate"), str(x.get("notes") or "")[:200],
        ])
    style_header(ws, 1); add_border(ws, 1, ws.max_row, 1, ws.max_column)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 24

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = os.path.join(OUTPUT_DIR, f"BizNext_Weekly_{datetime.today():%Y%m%d}.xlsx")
    wb.save(out)
    return out


def main() -> None:
    print("Fetching leads...", flush=True)
    leads = fetch_leads()
    print(f"Leads fetched: {len(leads)}", flush=True)
    today = datetime.today()
    out = build_workbook(leads, year=today.year, month=today.month)
    print(out)


if __name__ == "__main__":
    main()
